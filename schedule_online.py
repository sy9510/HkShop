from datetime import datetime, timedelta
import pytz
import requests
from loguru import logger
from retrying import retry
from openpyxl import load_workbook
import os


@retry(stop_max_attempt_number=3)
def safe_request(url, method, session=None, **kwargs):
    try:
        if session is None:
            session = requests.session()
        response = session.request(method, url, **kwargs)
        # logger.info(response.text)
        response.raise_for_status()
        return response
    except requests.RequestException as e:
        logger.error(f'Request to {url} failed: {str(e)}')
        raise


class BusinessPost:
    def __init__(self, app_id, app_secret, access_token):
        self.app_id = app_id
        self.app_secret = app_secret
        self.instagram_id_list = []
        self.page_id_list = []
        self.access_token = access_token
        self.refresh_token()
        self.init_account()

    def init_account(self):
        self.fetch_page_id_list()
        self.fetch_instagram_list()

    def fetch_page_id_list(self):
        url = f"https://graph.facebook.com/v18.0/me/accounts?access_token={self.access_token}"
        response = safe_request(url, 'GET')
        for shop_page in response.json()['data']:
            self.page_id_list.append(shop_page['id'])

    def fetch_instagram_list(self):
        for page_id in self.page_id_list:
            url = f"https://graph.facebook.com/v18.0/{page_id}?fields=instagram_business_account&access_token={self.access_token}"
            response = safe_request(url, 'GET')
            data = response.json()
            self.instagram_id_list.append(data.get('instagram_business_account').get('id'))

    def refresh_token(self):
        url = f"https://graph.facebook.com/v18.0/oauth/access_token"
        params = {
            'grant_type': 'fb_exchange_token',
            'client_id': self.app_id,
            'client_secret': self.app_secret,
            'fb_exchange_token': self.access_token
        }
        res = safe_request(url, 'GET', params=params)
        self.access_token = res.json().get('access_token')

    def fetch_page_access_token(self, page_id):
        url = f'https://graph.facebook.com/v18.0/me/accounts?access_token={self.access_token}'
        res = safe_request(url, 'GET')
        page_access_token = ""
        logger.info('Page access token was fetched successfully.')
        for item in res.json()['data']:
            if item['id'] == page_id:
                page_access_token = item['access_token']
                break
        return page_access_token

    @retry(stop_max_attempt_number=3)
    def upload_image(self, page_id, page_access_token, photo_path):
        try:
            url = f'https://graph.facebook.com/v18.0/{page_id}/photos?access_token={page_access_token}'
            files = {'file': open(photo_path, 'rb')}
            data = {'published': False}
            response = safe_request(url, 'POST', files=files, data=data)
            photo_id = response.json()['id']
            return photo_id
        except Exception as e:
            logger.error(f'Failed to upload image: {str(e)}')
            # Re-raise the exception if you want to let the caller handle it
            raise

    def post_facebook_page(self, page_id, page_access_token, message, scheduled_publish_time, photo_path,
                           photo_id_len=10):
        photo_file_list = []
        logger.info("Start to upload images.")
        for filename in os.listdir(photo_path):
            if filename.endswith(".jpg"):  # 确保只上传JPEG文件
                image_path = os.path.join(photo_path, filename)
                photo_file_list.append(image_path)
        logger.info(f"Total {len(photo_file_list)} images were found.")
        photo_id_list = []
        for photo_file in photo_file_list[:photo_id_len]:
            photo_id_list.append({'media_fbid': self.upload_image(page_id, page_access_token, photo_file)})

        logger.info(f"Total {len(photo_id_list)} images were uploaded.")

        url = f'https://graph.facebook.com/v18.0/{page_id}/feed?access_token={page_access_token}'
        data = {
            'message': message,
            'attached_media': photo_id_list,
            'scheduled_publish_time': scheduled_publish_time,
            'published': False
        }
        response = safe_request(url, 'POST', json=data)
        logger.info('Post was scheduled successfully.')
        return response.json()

    def post_instagram_page(self, instagram, message, photo_urls, photo_id_len=3):
        media_ids = []
        for photo_url in photo_urls:
            url = f'https://graph.facebook.com/v18.0/{instagram}/media'
            data = {
                'image_url': photo_url,
                'access_token': access_token
            }
            response = safe_request(url, 'POST', json=data)
            media_id = response.json()['id']
            media_ids.append(media_id)
        # 创建 Carousel 的媒体容器
        url = f'https://graph.facebook.com/v18.0/{instagram}/media'
        data = {
            'media_type': 'CAROUSEL',
            'children': ','.join(media_ids[:photo_id_len]),
            'caption': message,
            'access_token': access_token
        }
        response = safe_request(url, 'POST', json=data)
        carousel_id = response.json()['id']

        # 发布 Carousel
        url = f'https://graph.facebook.com/v18.0/{instagram}/media_publish'
        data = {'creation_id': carousel_id, 'access_token': access_token}
        response = safe_request(url, 'POST', json=data)
        if response.status_code == 200:
            logger.info(f"Instagram published successfully.")
        else:
            logger.info("Instagram published Failed.")


def schedule_post(app_id, app_secret, access_token):
    business_post = BusinessPost(app_id, app_secret, access_token)

    # 使用 openpyxl 读取 Excel 文件
    wb = load_workbook('schedule.xlsx')
    ws = wb.active

    # 获取当前的时间
    hk_tz = pytz.timezone('Asia/Hong_Kong')
    now = datetime.now(hk_tz)

    # 计算 30 分钟前的时间
    past = now - timedelta(minutes=30)

    # 计算 5 分钟后的时间
    future = now + timedelta(minutes=5)

    flag = 0

    business_post_instagram_list = set(business_post.instagram_id_list)
    # 遍历 Excel 文件中的所有行
    for row in ws.iter_rows(min_row=2):  # 假设第一行是标题行，所以从第二行开始
        # 获取行中的时间
        task_time = row[0].value  # 假设'規劃時間'在第一列
        task_time = hk_tz.localize(task_time)

        if past <= task_time <= future:
            # 判断是否为instagram账号
            instagram_list = set([col.strip() for col in row[1].value.split(',')])
            for instagram in list(business_post_instagram_list.intersection(instagram_list)):
                business_post.post_instagram_page(instagram, row[2].value,
                                                  row[3].value.split(','))  # 假设'內容'在第二列，'圖片列表'在第三列
                flag = 1

    if flag == 0:
        logger.info("No post was scheduled.")


if __name__ == '__main__':
    import os

    app_id = os.getenv('APP_ID')
    app_secret = os.getenv('APP_SECRET')
    access_token = os.getenv('ACCESS_TOKEN')
    schedule_post(app_id, app_secret, access_token)
    # business_post = BusinessPost(app_id, app_secret, access_token)
    # print(business_post.page_id_list)
    # print(business_post.instagram_id_list)
